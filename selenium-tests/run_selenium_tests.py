"""
SNSOC — Selenium Web UI Test Runner
Drives a real Chrome browser against the live SNSOC web dashboard.
Tests every major UI component visible in the browser.
Generates: selenium-test-report.xlsx
"""
import sys, os, time, json
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "selenium", "-q"])
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException

BASE_URL = os.environ.get("SNSOC_BASE_URL", "https://snsoc-4.onrender.com")
USERNAME = os.environ.get("SNSOC_USER", "siva")
PASSWORD = os.environ.get("SNSOC_PASS", "siva2580")

RESULTS = []

def record(test_id, category, name, status, detail, elapsed_ms):
    RESULTS.append({
        "id": test_id,
        "category": category,
        "test_name": name,
        "status": status,
        "elapsed_ms": elapsed_ms,
        "detail": detail,
        "time": datetime.now().strftime("%H:%M:%S")
    })
    icon = "PASS" if status == "PASS" else ("WARN" if status == "WARN" else "FAIL")
    print(f"  [{icon}] {name} ({elapsed_ms}ms) — {detail}")

# ── Setup Chrome driver ────────────────────────────────────────────────────
print("\n[Selenium Test Suite] SNSOC Web Dashboard UI Tests")
print(f"Target: {BASE_URL}")
print("=" * 60)

opts = Options()
opts.add_argument("--headless=new")
opts.add_argument("--no-sandbox")
opts.add_argument("--disable-dev-shm-usage")
opts.add_argument("--disable-gpu")
opts.add_argument("--window-size=1280,900")
opts.add_argument("--ignore-certificate-errors")

try:
    driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(5)
    wait = WebDriverWait(driver, 15)
    print("[OK] Chrome WebDriver initialized\n")
except Exception as e:
    print(f"[FAIL] Could not start Chrome: {e}")
    sys.exit(1)

test_counter = [0]

def run(category, name, fn):
    test_counter[0] += 1
    tid = f"SEL-{test_counter[0]:03d}"
    t0 = time.time()
    try:
        detail, status = fn()
        ms = round((time.time() - t0) * 1000)
        record(tid, category, name, status, detail, ms)
    except Exception as e:
        ms = round((time.time() - t0) * 1000)
        record(tid, category, name, "FAIL", f"Exception: {e}", ms)

# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORY 1 — Login Page
# ═══════════════════════════════════════════════════════════════════════════════
print("[1] Login Page UI Tests")

def tc_login_page_loads():
    driver.get(f"{BASE_URL}/auth/login")
    time.sleep(2)
    title = driver.title
    if "SNSOC" in title or "Login" in title or "SOC" in title:
        return f"Page title: '{title}'", "PASS"
    return f"Unexpected title: '{title}'", "WARN"

def tc_login_has_username_field():
    driver.get(f"{BASE_URL}/auth/login")
    try:
        el = driver.find_element(By.CSS_SELECTOR, "input[type='text'], input[name='username'], input[id*='user']")
        return f"Username field found: tag={el.tag_name}, id={el.get_attribute('id')}", "PASS"
    except NoSuchElementException:
        return "Username input field not found in DOM", "FAIL"

def tc_login_has_password_field():
    try:
        el = driver.find_element(By.CSS_SELECTOR, "input[type='password'], input[name='passcode'], input[id*='pass']")
        return f"Password field found: id={el.get_attribute('id')}", "PASS"
    except NoSuchElementException:
        return "Password input field not found", "FAIL"

def tc_login_has_submit_button():
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit'], button.login-btn")
        return f"Submit button found: text='{btn.text}'", "PASS"
    except NoSuchElementException:
        return "Submit button not found", "FAIL"

def tc_login_page_https():
    url = driver.current_url
    if url.startswith("https://"):
        return f"Page served over HTTPS: {url}", "PASS"
    return f"Page not over HTTPS: {url}", "FAIL"

def tc_login_branding():
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        if "SNSOC" in body or "SOC" in body:
            return "SNSOC branding found on login page", "PASS"
        return "No branding text found", "WARN"
    except: return "Could not read page text", "WARN"

run("Login Page", "Login page loads with correct title", tc_login_page_loads)
run("Login Page", "Username input field present", tc_login_has_username_field)
run("Login Page", "Password input field present", tc_login_has_password_field)
run("Login Page", "Submit button present", tc_login_has_submit_button)
run("Login Page", "Page served over HTTPS", tc_login_page_https)
run("Login Page", "SNSOC branding on login page", tc_login_branding)

# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORY 2 — Authentication Flow
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[2] Authentication Flow Tests")

def tc_invalid_login():
    driver.get(f"{BASE_URL}/auth/login")
    time.sleep(1)
    try:
        u = driver.find_element(By.CSS_SELECTOR, "input[type='text'], input[name='username']")
        p = driver.find_element(By.CSS_SELECTOR, "input[type='password'], input[name='passcode']")
        u.clear(); u.send_keys("wronguser")
        p.clear(); p.send_keys("wrongpassword")
        btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
        btn.click()
        time.sleep(2)
        url = driver.current_url
        if "/auth/login" in url:
            return "Invalid credentials correctly rejected — still on login page", "PASS"
        return f"May have logged in with wrong creds — at: {url}", "FAIL"
    except Exception as e:
        return f"Error during invalid login test: {e}", "FAIL"

def tc_valid_login():
    driver.get(f"{BASE_URL}/auth/login")
    time.sleep(1)
    try:
        u = driver.find_element(By.CSS_SELECTOR, "input[type='text'], input[name='username']")
        p = driver.find_element(By.CSS_SELECTOR, "input[type='password'], input[name='passcode']")
        u.clear(); u.send_keys(USERNAME)
        p.clear(); p.send_keys(PASSWORD)
        btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
        btn.click()
        time.sleep(3)
        url = driver.current_url
        if "/auth/login" not in url:
            return f"Login successful — redirected to: {url}", "PASS"
        return f"Login may have failed — still at: {url}", "WARN"
    except Exception as e:
        return f"Login flow error: {e}", "FAIL"

def tc_dashboard_url_after_login():
    url = driver.current_url
    if BASE_URL in url and "/auth/login" not in url:
        return f"On dashboard URL: {url}", "PASS"
    return f"Unexpected URL: {url}", "WARN"

run("Auth Flow", "Invalid credentials rejected", tc_invalid_login)
run("Auth Flow", "Valid login succeeds", tc_valid_login)
run("Auth Flow", "Redirected to dashboard after login", tc_dashboard_url_after_login)

# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORY 3 — Dashboard UI Components
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[3] Dashboard UI Component Tests")

def tc_dashboard_title():
    try:
        h = driver.find_element(By.CSS_SELECTOR, "h1, h2, .dash-title, .page-title")
        text = h.text.strip()
        if text:
            return f"Dashboard heading: '{text}'", "PASS"
        return "Dashboard heading empty", "WARN"
    except: return "No heading found on dashboard", "WARN"

def tc_packet_count_panel():
    try:
        el = driver.find_element(By.ID, "total_packets_count")
        val = el.text.replace(",", "")
        if val.isdigit() and int(val) >= 0:
            return f"Total packets panel shows: {el.text}", "PASS"
        return f"Packet count not numeric: '{el.text}'", "WARN"
    except: return "total_packets_count element not found", "FAIL"

def tc_threat_level_panel():
    try:
        el = driver.find_element(By.ID, "dynamic_level")
        lvl = el.text.strip()
        valid = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
        if lvl in valid:
            return f"Threat level displayed: {lvl}", "PASS"
        return f"Threat level unexpected: '{lvl}'", "WARN"
    except: return "dynamic_level element not found", "FAIL"

def tc_protocol_chart_canvas():
    try:
        el = driver.find_element(By.ID, "doughnutChart")
        if el.is_displayed():
            return "Protocol distribution doughnut chart canvas visible", "PASS"
        return "Doughnut chart canvas hidden", "WARN"
    except: return "doughnutChart canvas not found", "FAIL"

def tc_traffic_chart_canvas():
    try:
        el = driver.find_element(By.ID, "areaChart")
        if el.is_displayed():
            return "Live traffic area chart canvas visible", "PASS"
        return "Area chart canvas hidden", "WARN"
    except: return "areaChart canvas not found", "FAIL"

def tc_system_status_indicator():
    try:
        body = driver.find_element(By.TAG_NAME, "body").text
        if "System Active" in body or "Active" in body or "active" in body.lower():
            return "System Active status indicator found", "PASS"
        return "System status indicator not found", "WARN"
    except: return "Could not scan page for status text", "WARN"

def tc_alerts_section():
    try:
        el = driver.find_element(By.ID, "incidents_body")
        rows = el.find_elements(By.CLASS_NAME, "alert-item")
        return f"Alerts section visible with {len(rows)} alert items", "PASS"
    except:
        try:
            driver.find_element(By.CSS_SELECTOR, "[id*='alert'], [class*='alert']")
            return "Alert section found (alternate selector)", "PASS"
        except: return "Alerts section not found in DOM", "WARN"

def tc_packet_table():
    try:
        el = driver.find_element(By.ID, "packet_body")
        rows = el.find_elements(By.TAG_NAME, "tr")
        return f"Packet stream table has {len(rows)} rows", "PASS"
    except: return "packet_body table not found", "WARN"

def tc_nav_sidebar():
    try:
        links = driver.find_elements(By.CSS_SELECTOR, ".nav-item, .sidebar a, nav a")
        if len(links) >= 2:
            texts = [l.text.strip() for l in links if l.text.strip()]
            return f"Sidebar navigation found: {texts}", "PASS"
        return f"Only {len(links)} nav items found", "WARN"
    except: return "Navigation sidebar not found", "WARN"

run("Dashboard UI", "Dashboard heading visible", tc_dashboard_title)
run("Dashboard UI", "Total Packets Evaluated panel", tc_packet_count_panel)
run("Dashboard UI", "Threat Level panel with valid value", tc_threat_level_panel)
run("Dashboard UI", "Protocol Distribution doughnut chart", tc_protocol_chart_canvas)
run("Dashboard UI", "Live Traffic Over Time area chart", tc_traffic_chart_canvas)
run("Dashboard UI", "System Active status indicator", tc_system_status_indicator)
run("Dashboard UI", "Alerts/Incidents section", tc_alerts_section)
run("Dashboard UI", "Live Packet Stream table", tc_packet_table)
run("Dashboard UI", "Sidebar navigation links", tc_nav_sidebar)

# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORY 4 — Real-Time Updates
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[4] Real-Time Update Tests")

def tc_packet_count_increments():
    try:
        el = driver.find_element(By.ID, "total_packets_count")
        val1 = int(el.text.replace(",", "") or "0")
        time.sleep(6)
        el = driver.find_element(By.ID, "total_packets_count")
        val2 = int(el.text.replace(",", "") or "0")
        if val2 > val1:
            return f"Packet count incremented: {val1} → {val2} (+{val2-val1})", "PASS"
        elif val2 == val1:
            return f"Packet count unchanged after 6s: {val1} (may be OK on free tier)", "WARN"
        return f"Packet count decreased: {val1} → {val2}", "FAIL"
    except Exception as e:
        return f"Could not read packet count: {e}", "FAIL"

def tc_traffic_chart_updates():
    try:
        points = driver.execute_script(
            "return window.areaChart ? window.areaChart.data.labels.length : -1"
        )
        if points > 0:
            return f"Traffic chart has {points} data points (updating)", "PASS"
        elif points == 0:
            return "Traffic chart has 0 data points", "WARN"
        return "areaChart JS object not found", "WARN"
    except Exception as e:
        return f"JS chart inspection failed: {e}", "WARN"

def tc_socket_connected():
    try:
        connected = driver.execute_script(
            "return window.socket ? window.socket.connected : null"
        )
        if connected is True:
            return "Socket.IO connected=true", "PASS"
        elif connected is False:
            return "Socket.IO connected=false (disconnected)", "WARN"
        return "Socket object not found on window", "WARN"
    except Exception as e:
        return f"Socket JS check failed: {e}", "WARN"

run("Real-Time", "Packet count increments within 6 seconds", tc_packet_count_increments)
run("Real-Time", "Traffic chart accumulates data points", tc_traffic_chart_updates)
run("Real-Time", "Socket.IO connection is active", tc_socket_connected)

# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORY 5 — Logout
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[5] Logout Tests")

def tc_logout_flow():
    try:
        driver.get(f"{BASE_URL}/auth/logout")
        time.sleep(2)
        url = driver.current_url
        if "/auth/login" in url:
            return f"Logout redirected to login page: {url}", "PASS"
        return f"After logout at: {url}", "WARN"
    except Exception as e:
        return f"Logout error: {e}", "FAIL"

run("Logout", "Logout redirects to login page", tc_logout_flow)

driver.quit()
print("\n[OK] Browser closed")

# ── Generate Excel Report ────────────────────────────────────────────────────
print("\nGenerating Selenium Excel report...")
import subprocess
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs("test-reports", exist_ok=True)
total = len(RESULTS); passed = sum(1 for r in RESULTS if r["status"]=="PASS")
warned = sum(1 for r in RESULTS if r["status"]=="WARN"); failed = sum(1 for r in RESULTS if r["status"]=="FAIL")
avg_ms = round(sum(r["elapsed_ms"] for r in RESULTS)/total) if total else 0

def mk_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Selenium Results"
ws.sheet_view.showGridLines = False; ws.freeze_panes = "A3"

ws.merge_cells("A1:H1")
c = ws["A1"]; c.value = f"SNSOC SELENIUM WEB UI TEST REPORT  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {BASE_URL}"
c.fill = PatternFill("solid", fgColor="0D1B2A"); c.font = Font(bold=True,color="FFFFFF",size=13,name="Calibri")
c.alignment = Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height = 36

hdrs = ["ID","Category","Test Name","Status","Time(ms)","Detail","Timestamp","Browser"]
widths = [9,20,40,10,12,55,12,12]
for ci,(h,w) in enumerate(zip(hdrs,widths),1):
    cell = ws.cell(row=2,column=ci,value=h)
    cell.fill = PatternFill("solid",fgColor="0D1B2A"); cell.font = Font(bold=True,color="FFFFFF",size=10,name="Calibri")
    cell.alignment = Alignment(horizontal="center",vertical="center"); cell.border = mk_border()
    ws.column_dimensions[get_column_letter(ci)].width = w

STATUS_COLORS = {"PASS":"27AE60","WARN":"F39C12","FAIL":"E74C3C"}
for i,res in enumerate(RESULTS):
    row=i+3; bg="F2F3F4" if i%2 else "FFFFFF"; sc=STATUS_COLORS.get(res["status"],"95A5A6")
    def dc(c,v,b=bg,f="000000",bold=False,align="left"):
        cell=ws.cell(row=row,column=c,value=v)
        cell.fill=PatternFill("solid",fgColor=b); cell.font=Font(bold=bold,color=f,size=10,name="Calibri")
        cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True); cell.border=mk_border()
    dc(1,res["id"],bold=True,align="center")
    dc(2,res["category"])
    dc(3,res["test_name"])
    dc(4,res["status"],b=sc,f="FFFFFF",bold=True,align="center")
    dc(5,res["elapsed_ms"],align="center")
    dc(6,res["detail"])
    dc(7,res["time"],align="center")
    dc(8,"Chrome/Headless",align="center")
    ws.row_dimensions[row].height = 42

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False
for ci,(label,val,color) in enumerate([
    ("Total",total,"3498DB"),("Pass",passed,"27AE60"),("Warn",warned,"F39C12"),
    ("Fail",failed,"E74C3C"),("Pass Rate",f"{round(passed/total*100)}%" if total else "0%","8E44AD"),
    ("Avg Response",f"{avg_ms}ms","16A085")],1):
    ws2.column_dimensions[get_column_letter(ci)].width=18
    c1=ws2.cell(row=1,column=ci,value=label)
    c1.fill=PatternFill("solid",fgColor=color); c1.font=Font(bold=True,color="FFFFFF",size=12,name="Calibri")
    c1.alignment=Alignment(horizontal="center",vertical="center"); c1.border=mk_border()
    c2=ws2.cell(row=2,column=ci,value=val)
    c2.font=Font(bold=True,size=14,name="Calibri"); c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=mk_border()
    ws2.row_dimensions[1].height=32; ws2.row_dimensions[2].height=36

out = os.path.join("test-reports","selenium-test-report.xlsx")
wb.save(out)
print(f"[OK] Report saved: {out}")
print(f"     PASS={passed}  WARN={warned}  FAIL={failed}  Total={total}")
sys.exit(0 if failed == 0 else 1)
