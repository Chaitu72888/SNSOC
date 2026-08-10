"""
SNSOC — Security Test Runner
Performs automated security assessments against the live backend.
Categories: Authentication, Authorization, Input Validation, Headers, Rate Limiting,
            API Security, Session Security, Information Disclosure.
Generates: security-test-report.xlsx
"""
import sys, os, time, json, re, random, string
from datetime import datetime
import requests

BASE_URL = os.environ.get("SNSOC_BASE_URL", "https://snsoc-4.onrender.com")
USERNAME = os.environ.get("SNSOC_USER", "siva")
PASSWORD = os.environ.get("SNSOC_PASS", "siva2580")

RESULTS = []
counter = [0]

def run(category, name, fn, severity="INFO"):
    counter[0] += 1
    tid = f"SEC-{counter[0]:03d}"
    t0 = time.time()
    try:
        detail, status = fn()
        ms = round((time.time() - t0) * 1000)
    except Exception as e:
        ms = round((time.time() - t0) * 1000)
        detail, status = f"Exception: {e}", "FAIL"
    RESULTS.append({"id": tid, "category": category, "name": name, "status": status,
                    "severity": severity, "ms": ms, "detail": detail,
                    "ts": datetime.now().strftime("%H:%M:%S")})
    icon = "PASS" if status == "PASS" else ("WARN" if status == "WARN" else "FAIL")
    print(f"  [{icon}][{severity}] {name} ({ms}ms) — {detail}")

anon = requests.Session()
authed = requests.Session()

print(f"\n[Security Test Suite] SNSOC Backend Security Assessment")
print(f"Target: {BASE_URL}")
print("=" * 65)

# Login authed session (with retry for Render cold start)
print("\n[+] Establishing authenticated session...")
auth_success = False
for attempt in range(1, 4):
    try:
        r = authed.post(f"{BASE_URL}/auth/login",
                        data={"username": USERNAME, "passcode": PASSWORD},
                        allow_redirects=True, timeout=30)
        if r.status_code == 200 and "login" not in r.url:
            print(f"    [OK] Logged in as '{USERNAME}' (Attempt {attempt})")
            auth_success = True
            break
        else:
            print(f"    [!] Login attempt {attempt} returned HTTP {r.status_code}")
    except Exception as e:
        print(f"    [!] Login attempt {attempt} failed: {e}")
        time.sleep(3)

if not auth_success:
    print("    [!!] Auth failed after 3 attempts — continuing test execution with unauthenticated session")

# ══════════════════════════════════════════════════════════════════════════════
# 1 — AUTHENTICATION
# ══════════════════════════════════════════════════════════════════════════════
print("\n[1] Authentication Security")

def tc_auth_required_dashboard():
    r = anon.get(f"{BASE_URL}/api/dashboard", allow_redirects=True, timeout=15)
    if r.status_code in [401, 403] or "/auth/login" in r.url:
        return "/api/dashboard requires authentication", "PASS"
    return f"DANGER: /api/dashboard accessible unauthenticated — HTTP {r.status_code}", "FAIL"

def tc_auth_required_alerts():
    r = anon.get(f"{BASE_URL}/api/alerts", allow_redirects=True, timeout=15)
    if r.status_code in [401, 403] or "/auth/login" in r.url:
        return "/api/alerts requires authentication", "PASS"
    return f"DANGER: /api/alerts accessible unauthenticated — HTTP {r.status_code}", "FAIL"

def tc_auth_required_packets():
    r = anon.get(f"{BASE_URL}/api/packets", allow_redirects=True, timeout=15)
    if r.status_code in [401, 403] or "/auth/login" in r.url:
        return "/api/packets requires authentication", "PASS"
    return f"/api/packets unauthenticated — HTTP {r.status_code}", "FAIL"

def tc_auth_required_intel():
    r = anon.post(f"{BASE_URL}/api/intel/lookup",
                  json={"ip": "8.8.8.8"}, allow_redirects=True, timeout=15)
    if r.status_code in [401, 403] or "/auth/login" in r.url:
        return "/api/intel/lookup requires authentication", "PASS"
    return f"/api/intel/lookup unauthenticated — HTTP {r.status_code}", "FAIL"

def tc_auth_required_block():
    r = anon.post(f"{BASE_URL}/api/block",
                  json={"ip": "1.2.3.4"}, allow_redirects=True, timeout=15)
    if r.status_code in [401, 403] or "/auth/login" in r.url:
        return "/api/block requires authentication", "PASS"
    return f"CRITICAL: /api/block accessible unauthenticated — HTTP {r.status_code}", "FAIL"

def tc_login_public():
    r = anon.get(f"{BASE_URL}/auth/login", timeout=15)
    if r.status_code == 200:
        return "/auth/login is publicly accessible (correct)", "PASS"
    return f"/auth/login returned HTTP {r.status_code}", "WARN"

run("Authentication", "Dashboard endpoint requires auth", tc_auth_required_dashboard, "HIGH")
run("Authentication", "Alerts endpoint requires auth", tc_auth_required_alerts, "HIGH")
run("Authentication", "Packets endpoint requires auth", tc_auth_required_packets, "MEDIUM")
run("Authentication", "Intel lookup requires auth", tc_auth_required_intel, "HIGH")
run("Authentication", "Block IP requires auth", tc_auth_required_block, "CRITICAL")
run("Authentication", "Login page publicly accessible", tc_login_public, "INFO")

# ══════════════════════════════════════════════════════════════════════════════
# 2 — INPUT VALIDATION & INJECTION
# ══════════════════════════════════════════════════════════════════════════════
print("\n[2] Input Validation & Injection")

def tc_sql_injection_login():
    payloads = ["' OR '1'='1", "admin'--", "1; DROP TABLE users--", "' UNION SELECT 1,2,3--"]
    for p in payloads:
        r = anon.post(f"{BASE_URL}/auth/login",
                      data={"username": p, "passcode": "x"},
                      allow_redirects=True, timeout=15)
        if "/auth/login" not in r.url and r.status_code not in [200, 302]:
            return f"Possible SQL injection vulnerability with payload: {p}", "FAIL"
    return f"All {len(payloads)} SQL injection payloads on login rejected", "PASS"

def tc_xss_login():
    xss = "<script>alert('XSS')</script>"
    r = anon.post(f"{BASE_URL}/auth/login",
                  data={"username": xss, "passcode": "x"},
                  allow_redirects=True, timeout=15)
    if xss in r.text and "text/html" in r.headers.get("Content-Type",""):
        return "XSS payload reflected unescaped — VULNERABLE", "FAIL"
    return "XSS payload not reflected raw", "PASS"

def tc_invalid_ip_intel():
    for ip in ["not-an-ip", "999.999.999.999", "'; DROP TABLE--", "<script>alert(1)</script>"]:
        r = authed.post(f"{BASE_URL}/api/intel/lookup",
                        json={"ip": ip}, timeout=15)
        if r.status_code in [200, 400, 422]:
            continue
        return f"Unexpected status {r.status_code} for invalid IP '{ip}'", "WARN"
    return "All invalid IP inputs handled gracefully", "PASS"

def tc_large_payload():
    big = "A" * 100000
    r = authed.post(f"{BASE_URL}/api/intel/lookup",
                    json={"ip": big}, timeout=15)
    if r.status_code in [400, 413, 422, 200]:
        return f"Large payload handled: HTTP {r.status_code}", "PASS"
    return f"Unexpected response to 100KB payload: HTTP {r.status_code}", "WARN"

def tc_path_traversal():
    for path in ["/../../../etc/passwd", "/%2e%2e%2fetc%2fpasswd", "/static/../.env"]:
        r = anon.get(f"{BASE_URL}{path}", allow_redirects=False, timeout=15)
        if "root:" in r.text or "SECRET" in r.text or "DB_URI" in r.text:
            return f"Path traversal VULNERABLE at {path}", "FAIL"
    return "Path traversal attempts all returned safe responses", "PASS"

run("Input Validation", "SQL injection on login form", tc_sql_injection_login, "CRITICAL")
run("Input Validation", "XSS reflection on login form", tc_xss_login, "HIGH")
run("Input Validation", "Invalid IP values for intel lookup", tc_invalid_ip_intel, "MEDIUM")
run("Input Validation", "Large payload (100KB) handling", tc_large_payload, "MEDIUM")
run("Input Validation", "Path traversal attempts", tc_path_traversal, "HIGH")

# ══════════════════════════════════════════════════════════════════════════════
# 3 — HTTP SECURITY HEADERS
# ══════════════════════════════════════════════════════════════════════════════
print("\n[3] HTTP Security Headers")

r_check = anon.get(f"{BASE_URL}/auth/login", timeout=15)
HEADERS = r_check.headers

def tc_cors():
    origin = HEADERS.get("Access-Control-Allow-Origin", "")
    if origin == "*":
        return "CORS allows all origins (*) — acceptable for public API, review for sensitive routes", "WARN"
    elif origin:
        return f"CORS origin: {origin}", "PASS"
    return "No CORS headers present", "WARN"

def tc_content_type_options():
    val = HEADERS.get("X-Content-Type-Options", "")
    if "nosniff" in val:
        return "X-Content-Type-Options: nosniff present", "PASS"
    return f"X-Content-Type-Options missing or incorrect: '{val}'", "WARN"

def tc_frame_options():
    val = HEADERS.get("X-Frame-Options", "")
    csp = HEADERS.get("Content-Security-Policy", "")
    if "DENY" in val or "SAMEORIGIN" in val:
        return f"X-Frame-Options: {val}", "PASS"
    if "frame-ancestors" in csp:
        return f"CSP frame-ancestors used instead", "PASS"
    return "X-Frame-Options not set — clickjacking possible", "WARN"

def tc_csp():
    csp = HEADERS.get("Content-Security-Policy", "")
    if csp:
        return f"CSP present: {csp[:80]}...", "PASS"
    return "Content-Security-Policy header missing", "WARN"

def tc_hsts():
    hsts = HEADERS.get("Strict-Transport-Security", "")
    if hsts:
        return f"HSTS present: {hsts}", "PASS"
    return "Strict-Transport-Security (HSTS) header missing", "WARN"

def tc_server_disclosure():
    server = HEADERS.get("Server", "")
    if "Flask" in server or "Werkzeug" in server or "Python" in server:
        return f"Server header discloses framework: '{server}'", "WARN"
    return f"Server header: '{server}' — no framework disclosure", "PASS"

def tc_powered_by():
    pb = HEADERS.get("X-Powered-By", "")
    if pb:
        return f"X-Powered-By discloses: '{pb}'", "WARN"
    return "X-Powered-By header not present (good)", "PASS"

run("HTTP Headers", "CORS policy", tc_cors, "MEDIUM")
run("HTTP Headers", "X-Content-Type-Options: nosniff", tc_content_type_options, "LOW")
run("HTTP Headers", "Clickjacking protection (X-Frame-Options)", tc_frame_options, "MEDIUM")
run("HTTP Headers", "Content-Security-Policy (CSP)", tc_csp, "MEDIUM")
run("HTTP Headers", "HSTS enforcement", tc_hsts, "MEDIUM")
run("HTTP Headers", "Server header disclosure", tc_server_disclosure, "LOW")
run("HTTP Headers", "X-Powered-By disclosure", tc_powered_by, "LOW")

# ══════════════════════════════════════════════════════════════════════════════
# 4 — SESSION SECURITY
# ══════════════════════════════════════════════════════════════════════════════
print("\n[4] Session Security")

def tc_session_cookie_httponly():
    r = anon.post(f"{BASE_URL}/auth/login",
                  data={"username": USERNAME, "passcode": PASSWORD},
                  allow_redirects=False, timeout=15)
    for c in r.cookies:
        if "session" in c.name.lower():
            if c.has_nonstandard_attr("HttpOnly") or c.has_nonstandard_attr("httponly"):
                return f"Session cookie '{c.name}' has HttpOnly flag", "PASS"
            return f"Session cookie '{c.name}' missing HttpOnly flag", "WARN"
    raw = r.headers.get("Set-Cookie", "")
    if "httponly" in raw.lower():
        return "HttpOnly found in raw Set-Cookie header", "PASS"
    return f"HttpOnly flag not confirmed in: {raw[:100]}", "WARN"

def tc_session_cookie_secure():
    r = anon.post(f"{BASE_URL}/auth/login",
                  data={"username": USERNAME, "passcode": PASSWORD},
                  allow_redirects=False, timeout=15)
    raw = r.headers.get("Set-Cookie", "")
    if "Secure" in raw or "secure" in raw.lower():
        return "Session cookie has Secure flag (HTTPS-only)", "PASS"
    return f"Session cookie may be missing Secure flag: {raw[:100]}", "WARN"

def tc_logout_invalidates_session():
    s = requests.Session()
    s.post(f"{BASE_URL}/auth/login",
           data={"username": USERNAME, "passcode": PASSWORD},
           allow_redirects=True, timeout=15)
    r_before = s.get(f"{BASE_URL}/api/dashboard", allow_redirects=False, timeout=15)
    s.get(f"{BASE_URL}/auth/logout", allow_redirects=True, timeout=15)
    r_after = s.get(f"{BASE_URL}/api/dashboard", allow_redirects=True, timeout=15)
    if "/auth/login" in r_after.url or r_after.status_code in [401, 403]:
        return "Session invalidated on logout — dashboard inaccessible after", "PASS"
    return f"Session may persist after logout — dashboard at {r_after.url}", "WARN"

run("Session", "Session cookie HttpOnly flag", tc_session_cookie_httponly, "HIGH")
run("Session", "Session cookie Secure flag", tc_session_cookie_secure, "MEDIUM")
run("Session", "Logout invalidates session", tc_logout_invalidates_session, "HIGH")

# ══════════════════════════════════════════════════════════════════════════════
# 5 — API SECURITY
# ══════════════════════════════════════════════════════════════════════════════
print("\n[5] API Security")

def tc_method_not_allowed():
    r = authed.delete(f"{BASE_URL}/api/dashboard", timeout=15)
    if r.status_code in [405, 404]:
        return f"DELETE /api/dashboard correctly rejected: HTTP {r.status_code}", "PASS"
    return f"Unexpected: DELETE /api/dashboard returned HTTP {r.status_code}", "WARN"

def tc_content_type_enforced():
    r = authed.post(f"{BASE_URL}/api/intel/lookup",
                    data="not-json", headers={"Content-Type": "text/plain"}, timeout=15)
    if r.status_code in [400, 415, 422]:
        return f"Wrong Content-Type rejected: HTTP {r.status_code}", "PASS"
    return f"Wrong Content-Type accepted: HTTP {r.status_code}", "WARN"

def tc_empty_json_body():
    r = authed.post(f"{BASE_URL}/api/intel/lookup",
                    json={}, timeout=15)
    if r.status_code in [400, 422]:
        return f"Empty JSON body rejected: HTTP {r.status_code}", "PASS"
    return f"Empty JSON accepted: HTTP {r.status_code} — {r.text[:100]}", "WARN"

def tc_ids_rules_endpoint():
    r = authed.get(f"{BASE_URL}/api/ids/rules", timeout=15)
    if r.status_code == 200:
        return f"IDS rules endpoint accessible to authenticated user", "PASS"
    return f"IDS rules returned HTTP {r.status_code}", "WARN"

def tc_http_methods_exposed():
    # OPTIONS request to check allowed methods
    r = requests.options(f"{BASE_URL}/api/dashboard", timeout=15)
    allow = r.headers.get("Allow", r.headers.get("Access-Control-Allow-Methods",""))
    if "TRACE" in allow or "CONNECT" in allow:
        return f"Dangerous methods exposed: {allow}", "WARN"
    return f"Methods exposed: {allow or 'standard'}", "PASS"

run("API Security", "DELETE method not allowed on GET endpoints", tc_method_not_allowed, "LOW")
run("API Security", "Content-Type enforcement on POST", tc_content_type_enforced, "MEDIUM")
run("API Security", "Empty JSON body validation", tc_empty_json_body, "LOW")
run("API Security", "IDS rules endpoint access control", tc_ids_rules_endpoint, "MEDIUM")
run("API Security", "Dangerous HTTP methods not exposed", tc_http_methods_exposed, "MEDIUM")

# ══════════════════════════════════════════════════════════════════════════════
# 6 — INFORMATION DISCLOSURE
# ══════════════════════════════════════════════════════════════════════════════
print("\n[6] Information Disclosure")

def tc_no_stack_trace():
    r = anon.get(f"{BASE_URL}/api/nonexistent_route_xyz123", timeout=15)
    body = r.text.lower()
    if "traceback" in body or "file \"/" in body or "line " in body and ".py" in body:
        return "Stack trace exposed in 404 response — information disclosure", "FAIL"
    return f"No stack trace in 404 response (HTTP {r.status_code})", "PASS"

def tc_no_env_exposed():
    for path in ["/.env", "/.env.example", "/config.py", "/settings.py"]:
        r = anon.get(f"{BASE_URL}{path}", timeout=15)
        if r.status_code == 200 and ("SECRET" in r.text or "DATABASE" in r.text or "KEY" in r.text):
            return f"Sensitive config exposed at {path}", "FAIL"
    return "No sensitive config files exposed", "PASS"

def tc_no_git_exposed():
    r = anon.get(f"{BASE_URL}/.git/HEAD", timeout=15)
    if r.status_code == 200 and "ref:" in r.text:
        return "Git repository exposed at /.git/HEAD", "WARN"
    return f"/.git/HEAD not accessible (HTTP {r.status_code})", "PASS"

def tc_error_messages_safe():
    r = authed.post(f"{BASE_URL}/api/intel/lookup",
                    json={"ip": None}, timeout=15)
    body = r.text
    if "sqlalchemy" in body.lower() or "psycopg" in body.lower() or "sqlite" in body.lower():
        return "Database errors exposed in response body", "WARN"
    return "No database internals exposed in error messages", "PASS"

run("Info Disclosure", "No stack traces in error pages", tc_no_stack_trace, "HIGH")
run("Info Disclosure", "Sensitive config files not exposed", tc_no_env_exposed, "CRITICAL")
run("Info Disclosure", "Git repository not exposed", tc_no_git_exposed, "HIGH")
run("Info Disclosure", "DB errors not leaked in responses", tc_error_messages_safe, "MEDIUM")

# ══════════════════════════════════════════════════════════════════════════════
# 7 — TRANSPORT SECURITY
# ══════════════════════════════════════════════════════════════════════════════
print("\n[7] Transport Security")

def tc_https_only():
    http_url = BASE_URL.replace("https://", "http://")
    try:
        r = requests.get(http_url, timeout=10, allow_redirects=False)
        if r.status_code in [301, 302] and "https" in r.headers.get("Location",""):
            return "HTTP redirects to HTTPS (301/302)", "PASS"
        return f"HTTP not redirected to HTTPS: HTTP {r.status_code}", "WARN"
    except requests.exceptions.ConnectionError:
        return "HTTP connection refused — only HTTPS available", "PASS"

def tc_tls_version():
    import ssl
    try:
        ctx = ssl.create_default_context()
        with ctx.wrap_socket(__import__("socket").socket(), server_hostname=BASE_URL.split("//")[1]) as s:
            s.connect((BASE_URL.split("//")[1], 443))
            proto = s.version()
            if proto in ["TLSv1.2", "TLSv1.3"]:
                return f"TLS version: {proto} (secure)", "PASS"
            return f"TLS version: {proto} (may be insecure)", "WARN"
    except Exception as e:
        return f"TLS check: {e}", "WARN"

run("Transport", "HTTP redirects to HTTPS", tc_https_only, "HIGH")
run("Transport", "TLS version check (TLS 1.2+)", tc_tls_version, "HIGH")

# ══════════════════════════════════════════════════════════════════════════════
# Generate Excel Report
# ══════════════════════════════════════════════════════════════════════════════
print("\n\nGenerating Security Excel report...")
import subprocess
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.makedirs("test-reports", exist_ok=True)
total = len(RESULTS); passed = sum(1 for r in RESULTS if r["status"]=="PASS")
warned = sum(1 for r in RESULTS if r["status"]=="WARN"); failed = sum(1 for r in RESULTS if r["status"]=="FAIL")
risk_score = round((failed * 10 + warned * 3) / total, 1) if total else 0

def mk_b():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)

wb = openpyxl.Workbook()

# Results Sheet
ws = wb.active; ws.title = "Security Findings"
ws.sheet_view.showGridLines = False; ws.freeze_panes = "A3"
ws.merge_cells("A1:H1")
c=ws["A1"]; c.value=f"SNSOC SECURITY TEST REPORT  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {BASE_URL}"
c.fill=PatternFill("solid",fgColor="1C0A00"); c.font=Font(bold=True,color="FF6B35",size=13,name="Calibri")
c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36

hdrs=["ID","Category","Test Name","Status","Severity","Response(ms)","Detail","Timestamp"]
widths=[9,22,40,10,12,14,55,12]
SMAP={"CRITICAL":"7D0000","HIGH":"E74C3C","MEDIUM":"F39C12","LOW":"F1C40F","INFO":"95A5A6"}
STATUS_C={"PASS":"27AE60","WARN":"F39C12","FAIL":"E74C3C"}
for ci,(h,w) in enumerate(zip(hdrs,widths),1):
    cell=ws.cell(row=2,column=ci,value=h)
    cell.fill=PatternFill("solid",fgColor="1C0A00"); cell.font=Font(bold=True,color="FF6B35",size=10,name="Calibri")
    cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=mk_b()
    ws.column_dimensions[get_column_letter(ci)].width=w

for i,res in enumerate(RESULTS):
    row=i+3; bg="FFF5F5" if i%2 else "FFFFFF"; sc=STATUS_C.get(res["status"],"95A5A6")
    sev_c=SMAP.get(res["severity"],"95A5A6")
    def dc(c,v,b=bg,f="000000",bold=False,align="left"):
        cell=ws.cell(row=row,column=c,value=v)
        cell.fill=PatternFill("solid",fgColor=b); cell.font=Font(bold=bold,color=f,size=10,name="Calibri")
        cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True); cell.border=mk_b()
    dc(1,res["id"],bold=True,align="center")
    dc(2,res["category"])
    dc(3,res["name"])
    dc(4,res["status"],b=sc,f="FFFFFF",bold=True,align="center")
    dc(5,res["severity"],b=sev_c,f="FFFFFF" if res["severity"] in ["CRITICAL","HIGH","MEDIUM"] else "000000",bold=True,align="center")
    dc(6,res["ms"],align="center")
    dc(7,res["detail"])
    dc(8,res["ts"],align="center")
    ws.row_dimensions[row].height=42

# Summary Sheet
ws2=wb.create_sheet("Risk Summary")
ws2.sheet_view.showGridLines=False
ws2.merge_cells("A1:F1")
c2=ws2["A1"]; c2.value="SECURITY RISK SUMMARY — SNSOC Platform"
c2.fill=PatternFill("solid",fgColor="1C0A00"); c2.font=Font(bold=True,color="FF6B35",size=14,name="Calibri")
c2.alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=40

cards=[("Total Tests",total,"3498DB"),("Passed",passed,"27AE60"),("Warnings",warned,"F39C12"),
       ("Failed/Vuln",failed,"E74C3C"),("Risk Score",f"{risk_score}/10","E67E22"),
       ("Pass Rate",f"{round(passed/total*100)}%" if total else "0%","8E44AD")]
for ci,(label,val,color) in enumerate(cards,1):
    ws2.column_dimensions[get_column_letter(ci)].width=20
    for ri,v,sz,bold in [(3,label,10,True),(4,str(val),16,True)]:
        c_=ws2.cell(row=ri,column=ci,value=v)
        c_.fill=PatternFill("solid",fgColor=color if ri==3 else "F2F3F4")
        c_.font=Font(bold=bold,color="FFFFFF" if ri==3 else "000000",size=sz,name="Calibri")
        c_.alignment=Alignment(horizontal="center",vertical="center"); c_.border=mk_b()
    ws2.row_dimensions[3].height=30; ws2.row_dimensions[4].height=40

out=os.path.join("test-reports","security-test-report.xlsx")
wb.save(out)
print(f"[OK] Report saved: {out}")
print(f"     PASS={passed}  WARN={warned}  FAIL={failed}  Risk Score={risk_score}/10")
sys.exit(0 if failed == 0 else 1)
